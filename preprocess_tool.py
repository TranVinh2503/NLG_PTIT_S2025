import yaml
import random
import unicodedata
import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer
import openpyxl
import json

list_text = [
        "Opps, Tôi sẽ trả lời bạn vấn đề này sớm",
        "Vấn đề này tôi vẫn đang suy nghĩ , tôi sẽ trả lời bạn sau nhé.",
        "Để thêm câu trả lời về vấn đề này hãy liên hệ: 1800.599.989"]

def generate_default_responses():
    text = random.choice(list_text)
    return [{'text': text}]

def auto_fill_actions(stories_file, domain_file):
    # Load stories.yml
    with open(stories_file, 'r', encoding='utf-8') as f:
        stories_data = yaml.safe_load(f)

    # Load domain.yml
    with open(domain_file, 'r', encoding='utf-8') as f:
        domain_data = yaml.safe_load(f)

    # Extract actions from stories.yml
    actions = set()
    for story in stories_data.get('stories', []):
        for step in story.get('steps', []):
            if 'action' in step:
                actions.add(step['action'])

    # Update responses in domain.yml
    for action in actions:
        if action not in domain_data['responses']:
            domain_data['responses'][action] = generate_default_responses()
        else:
            domain_data['responses'][action].append({'text': generate_default_responses()[0]['text']})

    # Write updated domain.yml
    with open(domain_file, 'w', encoding='utf-8') as f:
        yaml.dump(domain_data, f, default_flow_style=False, allow_unicode=True)

def intents_miss(nlu_file,domain_file):
    # Load nlu.yml
    with open(nlu_file, 'r', encoding='utf-8') as f:
        nlu_data = yaml.safe_load(f)

    # Load new domain data
    with open(domain_file, 'r', encoding='utf-8') as f:
        domain_data = yaml.safe_load(f)

    
    nlu_intents = nlu_data.get('nlu', [])
    existing_intents = domain_data.get('intents', [])

    intents_miss = []
    for intent in nlu_intents:
        if intent not in existing_intents:
            intents_miss.append(intent.get('intent'))
    
    return intents_miss

def merge_domain(n_domain_file, t_domain_file):
    # Load domain.yml
    with open(t_domain_file, 'r', encoding='utf-8') as f:
        domain_data = yaml.safe_load(f)

    # Load new domain data
    with open(n_domain_file, 'r', encoding='utf-8') as f:
        new_domain_data = yaml.safe_load(f)

    # Merge intents
    new_intents = new_domain_data.get('intents', [])
    existing_intents = domain_data.get('intents', [])
    for intent in new_intents:
        if intent not in existing_intents:
            existing_intents.append(intent)

    intents_miss_array = intents_miss('data/nlu/teaching_methods.yml',t_domain_file)
    for intent in intents_miss_array:
        if intent not in existing_intents:
            existing_intents.append(intent)

    # Merge actions
    new_actions = new_domain_data.get('actions', [])
    existing_actions = domain_data.get('actions', [])
    for action in new_actions:
        if action not in existing_actions:
            existing_actions.append(action)

    # Merge responses
    new_responses = new_domain_data.get('responses', {})
    existing_responses = domain_data.get('responses', {})
    for key, value in new_responses.items():
        if key not in existing_responses:
            existing_responses[key] = value

    # Update domain data
    domain_data['intents'] = existing_intents
    domain_data['actions'] = existing_actions
    domain_data['responses'] = existing_responses

    # Write back to domain.yml
    with open(t_domain_file, 'w', encoding='utf-8') as f:
        yaml.dump(domain_data, f, default_flow_style=False, allow_unicode=True)

    print("Domain data merged successfully.")

"""Checks whether the content of an action is missing or empty."""
def check_miss_action_content(domain_file):
    with open(domain_file, 'r', encoding='utf-8') as f:
        domain_data = yaml.safe_load(f)
    
    actions = domain_data['responses']

    missing_or_empty_actions = []
    for action in actions:
        dict_elements = [item for item in actions[action] if isinstance(item, dict)]
        
        for dictionary in dict_elements:
            if(dictionary['text'] == '' or dictionary['text'] in list_text):
                missing_or_empty_actions.append(action)
    
    return missing_or_empty_actions
def remove_accents(text):
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

# The function converts the name to the "utter_" prefix and unsigned lower case
def format_name(text):
    return "utter_" + remove_accents(text).replace(' ', '_').lower()

def convert_responses(txt_file,domain_file):
    with open(txt_file, 'r', encoding='utf-8') as file:
        content = file.read()

    sections = [section.strip() for section in content.split('\n\n') if section.strip()]

    
    responses = {}
    for i, section in enumerate(sections, start=1):
        section_name = format_name(section.split('\n')[0])
        responses[section_name] = [{'text': '\n'.join(section.split('\n')[1:])}]

    yml_file = domain_file
    # export to yml file
    with open(yml_file, 'w', encoding='utf-8') as file:
        yaml.dump({'responses': responses}, file, default_flow_style=False, allow_unicode=True)

def check_number_actions(actions_file):
    with open(actions_file, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)

    actions = data['responses']
    print(len(actions))

def check_number_intents(nlu_file):
    with open(nlu_file, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)

    intents = data['nlu']
    print(len(intents))

def write_stories_file(domain_file,stories_file):
    with open(domain_file,'r', encoding="utf-8") as f:
        domain_file  =  yaml.safe_load(f)

    with open(stories_file,'w',encoding='utf-8') as f:
        f.write('stories:\n')
        for intent in domain_file["responses"]:
            nlu_intent = intent.replace("utter_","")
            f.write(f"- story: {intent}\n")
            f.write("  steps:\n")
            f.write(f"  - intent: {nlu_intent}\n")
            f.write(f"  - action: {intent}\n\n")
        f.close()

def auto_fill_actions_intents(nlu_file,stories_file,domain_file):
    with open(nlu_file, "r", encoding="utf-8") as file:
        nlu_data = yaml.safe_load(file)

    try:
        with open(stories_file, "r", encoding="utf-8") as file:
            stories_data = yaml.safe_load(file)
    except FileNotFoundError:
        stories_data = {}
            
    # Get unique intents from nlu.yml
    intents = set()
    for item in nlu_data["nlu"]:
        intents.add(item["intent"])
     # Read existing content of domain.yml
    try:
        with open(domain_file, "r", encoding="utf-8") as file:
            existing_data = yaml.safe_load(file)
    except FileNotFoundError:
        existing_data = {}


    #Get unique actions from stories.yml
    actions = set()
    for story in stories_data["stories"]:
        for step in story["steps"]:
            if "action" in step:
                actions.add(step["action"])

    # Update domain.yml
    if existing_data:
        existing_data["intents"] = list(intents.union(existing_data.get("intents", [])))
        existing_data["actions"] = list(actions.union(existing_data.get("actions", [])))
    else:
        existing_data = {
            "intents": list(intents),
            "actions": list(actions),
            "entities": []  # Add entities here if collected from nlu.yml
        }

    # Write domain.yml
    with open(domain_file, "w", encoding="utf-8") as file:
        yaml.dump(existing_data, file, allow_unicode=True)

def remove_stopwords(text):
    # Khởi tạo Porter Stemmer
    ps = PorterStemmer()
    question = text
    # Tokenization: chia câu hỏi thành các từ
    words = word_tokenize(question)
    # Loại bỏ stop words
    stop_words = set(stopwords.words("vietnamese"))
    filtered_question = [word for word in words if word.casefold() not in stop_words and word != '?']

    # Stemming: chuyển từ về dạng gốc
    stemmed_question = [ps.stem(word) for word in filtered_question]

    # Tạo cụm từ chỉ nội dung
    content_cluster = " ".join(stemmed_question)

    return content_cluster


def get_question_excel(excel_file, question_col_number, answer_col_number):
    file_path = excel_file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    column_D_values = []
    column_E_values = []

    # Lặp qua từng hàng trong cột D và lưu giá trị vào list
    for row in sheet.iter_rows(min_row=4, max_row=110, min_col=question_col_number, max_col=question_col_number, values_only=True):
        for cell in row:
            column_D_values.append(cell)

    # Lặp qua từng hàng trong cột D và lưu giá trị vào list
    for row in sheet.iter_rows(min_row=4, max_row=110, min_col=answer_col_number, max_col=answer_col_number, values_only=True):
        for cell in row:
            column_E_values.append(cell)

    formatted_question = []
    for question in column_D_values:
        question = remove_stopwords(question)
        question = format_name(question)
        question = question.replace('/','_')
        formatted_question.append(question)

    
    with open('domain_collect/common_question_v2.yml', 'w', encoding="utf-8") as f:
        f.write("responses:\n")
        for question, answer in zip(formatted_question, column_E_values):
            # Tạo tên hàm cắt nghĩa câu hỏi (ví dụ: utter_greet)
            function_name = question
            # Viết dữ liệu vào file
            f.write(f"  {function_name}:\n")
            f.write(f"  - text: '{answer}'\n")
    
    with open('data/nlu/ommon_question_v2.yml', 'w', encoding="utf-8") as f:
        f.write("nlu:\n")
        for intent,question in zip(formatted_question,column_D_values):
            intent_name = intent.replace('utter_', '')
            f.write(f'- intent: {intent_name}\n')
            f.write('  examples: |\n')
            f.write(f'    - {question}\n')
        f.write('\n')

    # Tạo file story.yml
    with open('data/stories/ommon_question_v2.yml', 'w', encoding="utf-8") as file:
        file.write(f'stories:\n')
        for question, response in zip(column_D_values, formatted_question):
            file.write(f"- story: {remove_stopwords(question)}\n")
            file.write("  steps:\n")

            # Tạo intent
            intent_name = response.replace('utter_', '')
            file.write(f"  - intent: {intent_name}\n")

            # Tạo response
            file.write(f"  - action: {response}\n")
            file.write('\n')

def add_suffix_course_name(domain_file,text):
     # Read existing content of domain.yml
    try:
        with open(domain_file, "r", encoding="utf-8") as file:
            existing_data = yaml.safe_load(file)
    except FileNotFoundError:
        existing_data = {}

    # Create a list of keys to iterate over
    action_names = list(existing_data["responses"].keys())

    # Iterate over the list of keys and add suffix to each action name
    for action_name in action_names:
        if not action_name.endswith(text):
            action_content = existing_data["responses"][action_name]
            existing_data["responses"][action_name + text] = action_content
            del existing_data["responses"][action_name]

    # Write updated content back to domain.yml
    with open(domain_file, "w", encoding="utf-8") as file:
        yaml.dump(existing_data, file, allow_unicode=True)

def convert_json_to_txt(json_file_path, txt_file_path):
    with open(json_file_path, 'r', encoding="utf-8") as f:
        data = json.load(f)

    with open(txt_file_path, 'a', encoding="utf-8") as f:
        f.write('\n')
        for item in data:
            for key, value in item.items():
                f.write(f"{key}: {value}\n")

    with open(txt_file_path,'r',encoding='utf-8') as f:
        lines = f.readlines()
    
    with open(txt_file_path,'w',encoding = 'utf-8') as f:
        for line in lines:
            if line.startswith("STT: 1"):
                continue
            if line.startswith("STT"):
                f.write("\n")
                continue
            if line.startswith("UNIT"):
                f.write(line)
            f.write(line)


def main():
    """If u have excel file which contains set of questions. You can run this function here to auto create file domain particular, stories, and nlu file. Function has 3 parameters. there is  path of excel file, number of column,which contains question and answer."""
    # get_question_excel('domain_collect/common-question_v2.xlsx',1,2)
    

    """Step 1: Convert  JSON to TXT file which converts to domain.yml file """
    # convert_json_to_txt('domain_collect/json/happy_kids_cam3b.json', 'domain_collect/text/happy_kids_cam3b.txt')

    """Step 2: Convert responses from txt file"""
    # convert_responses("domain_collect/text/happy_kids_cam3b.txt","domain_collect/happy_kids_cam3b.yml")
    
    """Step 3: Add suffix course name to clear information about data"""
    # add_suffix_course_name("domain_collect/happy_kids_cam3b.yml",'_happy_kids_cam3b')

    """Step 4: Auto created stories file by the intent and actions from domain file"""
    # write_stories_file('domain_collect/happy_kids_cam3b.yml','data/stories/happy_kids_cam3b.yml')

    """Step 5: Use chatbot to generate question in nlu file"""

    """Step 6: Auto fill actions and intents in the head of particular domain file """
    auto_fill_actions_intents('data/nlu/common_question.yml','data/stories/common_question.yml','domain.yml')   

    """Step 7: Merge particular domain to total domain( domain in root)"""
    # merge_domain('domain_collect/happy_kids_cam3b.yml','domain.yml')


if __name__ == '__main__':
    main()
